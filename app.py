from flask import Flask, render_template, request, redirect, session, send_file, url_for, flash
import io
from datetime import datetime
import pymysql
from werkzeug.security import check_password_hash, generate_password_hash
from functools import wraps
from config import Config
from zoneinfo import ZoneInfo
import secrets
from datetime import timedelta
import os
from werkzeug.utils import secure_filename
import random

# ==========================================
# REPORTLAB PARA PDF
# ==========================================
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

# ==========================================
# FLASK-MAIL PARA CORREOS
# ==========================================
from flask_mail import Mail, Message

# ==========================================
# DOTENV PARA VARIABLES DE ENTORNO
# ==========================================
from dotenv import load_dotenv
load_dotenv()

# ==========================
# APP
# ==========================
app = Flask(__name__)
app.config.from_object(Config)
print("🔍 CONFIGURACIÓN DE BASE DE DATOS:")
print(f"  HOST: {app.config['MYSQL_HOST']}")
print(f"  PORT: {app.config['MYSQL_PORT']}")
print(f"  USER: {app.config['MYSQL_USER']}")
print(f"  DB: {app.config['MYSQL_DB']}")
print(f"  PASSWORD: {'*' * len(app.config['MYSQL_PASSWORD'])}")
app.secret_key = app.config["SECRET_KEY"]

# ==========================
# SEGURIDAD DE SESIONES
# ==========================
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(minutes=30)
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = True  # Solo HTTPS en producción

# ==========================
# CONFIGURACIÓN DE CORREO
# ==========================
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USE_SSL'] = False
app.config['MAIL_USERNAME'] = 'docoutofernandodaniel@gmail.com'
app.config['MAIL_PASSWORD'] = 'xcxk gyhd aupy niwx'  
app.config['MAIL_DEFAULT_SENDER'] = 'docoutofernandodaniel@gmail.com'

# ==========================
# INICIALIZAR FLASK-MAIL
# ==========================
mail = Mail(app)

# ==========================
# CONFIGURACIÓN DE SUBIDA DE ARCHIVOS
# ==========================
UPLOAD_FOLDER = os.path.join('static', 'uploads')
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'doc', 'docx'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ==========================
# CONEXIÓN MYSQL
# ==========================
def obtener_conexion():
    try:
        conexion = pymysql.connect(
            host=app.config["MYSQL_HOST"],
            port=app.config["MYSQL_PORT"],
            user=app.config["MYSQL_USER"],
            password=app.config["MYSQL_PASSWORD"],
            database=app.config["MYSQL_DB"],
            charset=app.config["MYSQL_CHARSET"],
            cursorclass=pymysql.cursors.DictCursor,
            connect_timeout=10
        )
        conexion.ping(reconnect=True)
        print("✅ MYSQL RAILWAY CONECTADO")
        return conexion
    except Exception as error:
        print("❌ ERROR MYSQL:")
        print(error)
        return None

# ==========================
# FECHA Y HORA VENEZUELA
# ==========================
def obtener_fecha_venezuela():
    return datetime.now(ZoneInfo("America/Caracas"))

# ==========================================
# AUDITORÍA DEL SISTEMA
# ==========================================
def registrar_auditoria(accion, modulo):
    """Registra una acción en la tabla auditoria y devuelve el ID insertado."""
    if "id_usuario" not in session:
        return None
    conexion = obtener_conexion()
    if conexion is None:
        return None
    cursor = conexion.cursor()
    ip = request.remote_addr if request else "0.0.0.0"
    try:
        cursor.execute("""
            INSERT INTO auditoria
            (id_usuario, usuario, rol, accion, modulo, fecha, ip)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (
            session.get("id_usuario"),
            session.get("usuario"),
            session.get("rol"),
            accion,
            modulo,
            obtener_fecha_venezuela(),
            ip
        ))
        conexion.commit()
        id_auditoria = cursor.lastrowid
        return id_auditoria
    except Exception as e:
        print(f"❌ Error al registrar auditoría: {e}")
        return None
    finally:
        cursor.close()
        conexion.close() 
        
# ==========================================
# REGISTRAR DETALLE DE AUDITORÍA
# ==========================================
def registrar_detalle_auditoria(id_auditoria, campo, valor_anterior, valor_nuevo):
    """Registra un cambio específico en la tabla detalle_auditoria"""
    if not id_auditoria:
        return
    conexion = obtener_conexion()
    if conexion is None:
        return
    cursor = conexion.cursor()
    try:
        cursor.execute("""
            INSERT INTO detalle_auditoria
            (id_auditoria, campo, valor_anterior, valor_nuevo)
            VALUES (%s, %s, %s, %s)
        """, (id_auditoria, campo, valor_anterior, valor_nuevo))
        conexion.commit()
    except Exception as e:
        print(f"❌ Error al registrar detalle de auditoría: {e}")
    finally:
        cursor.close()
        conexion.close()  

# ==========================================
# FUNCIÓN PARA ENVIAR CORREOS (UNIFICADA)
# ==========================================
def enviar_correo(destinatario, asunto, mensaje_html):
    try:
        msg = Message(
            subject=asunto,
            recipients=[destinatario],
            html=mensaje_html,
            sender=app.config['MAIL_DEFAULT_SENDER']
        )
        mail.send(msg)
        print(f"✅ Correo enviado a {destinatario}")
        if "id_usuario" in session:
            registrar_auditoria(
                f"Correo enviado a {destinatario}",
                "Notificaciones"
            )
        return True
    except Exception as e:
        print(f"❌ Error al enviar correo: {e}")
        import traceback
        traceback.print_exc()
        return False

# ==========================
# DECORADORES
# ==========================
def login_requerido(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "usuario" not in session:
            return redirect("/login")
        return f(*args, **kwargs)
    return decorated_function

def rol_requerido(*roles):
    def decorador(f):
        @wraps(f)
        def funcion_protegida(*args, **kwargs):
            if "rol" not in session:
                return redirect("/login")
            if session["rol"] not in roles:
                return "⛔ Acceso no autorizado.", 403
            return f(*args, **kwargs)
        return funcion_protegida
    return decorador

# ============================================
# FUNCIÓN PARA CREAR TABLAS
# ============================================
def crear_tablas():
    """Crea todas las tablas si no existen"""
    conexion = obtener_conexion()
    if conexion is None:
        print("❌ No se pudo conectar para crear tablas")
        return
    
    cursor = conexion.cursor()
    
    sql_script = """
    CREATE TABLE IF NOT EXISTS usuarios (
        id_usuario INT AUTO_INCREMENT PRIMARY KEY,
        usuario VARCHAR(50) UNIQUE NOT NULL,
        password VARCHAR(255) NOT NULL,
        rol ENUM('administrador', 'docente', 'estudiante') NOT NULL,
        email VARCHAR(100),
        fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );
    
    CREATE TABLE IF NOT EXISTS estudiantes (
        id_estudiante INT AUTO_INCREMENT PRIMARY KEY,
        id_usuario INT NOT NULL,
        cedula VARCHAR(20) UNIQUE NOT NULL,
        nombres VARCHAR(100) NOT NULL,
        apellidos VARCHAR(100) NOT NULL,
        carrera VARCHAR(100),
        semestre INT,
        FOREIGN KEY (id_usuario) REFERENCES usuarios(id_usuario) ON DELETE CASCADE
    );
    
    CREATE TABLE IF NOT EXISTS docentes (
        id_docente INT AUTO_INCREMENT PRIMARY KEY,
        id_usuario INT NOT NULL,
        cedula VARCHAR(20) UNIQUE NOT NULL,
        nombres VARCHAR(100) NOT NULL,
        apellidos VARCHAR(100) NOT NULL,
        departamento VARCHAR(100),
        FOREIGN KEY (id_usuario) REFERENCES usuarios(id_usuario) ON DELETE CASCADE
    );
    
    CREATE TABLE IF NOT EXISTS solicitudes (
        id_solicitud INT AUTO_INCREMENT PRIMARY KEY,
        id_usuario INT NOT NULL,
        fecha DATE NOT NULL,
        tipo_solicitud VARCHAR(100) NOT NULL,
        descripcion TEXT,
        estado ENUM('Pendiente', 'Aprobada', 'Rechazada') DEFAULT 'Pendiente',
        fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (id_usuario) REFERENCES usuarios(id_usuario) ON DELETE CASCADE
    );
    
    CREATE TABLE IF NOT EXISTS beneficios (
        id_beneficio INT AUTO_INCREMENT PRIMARY KEY,
        id_estudiante INT NOT NULL,
        nombre_beneficio VARCHAR(100) NOT NULL,
        descripcion TEXT,
        fecha_solicitud DATE NOT NULL,
        estado ENUM('Pendiente', 'Aprobado', 'Rechazado') DEFAULT 'Pendiente',
        FOREIGN KEY (id_estudiante) REFERENCES estudiantes(id_estudiante) ON DELETE CASCADE
    );
    
    CREATE TABLE IF NOT EXISTS evaluaciones (
        id_evaluacion INT AUTO_INCREMENT PRIMARY KEY,
        id_estudiante INT NOT NULL,
        asignatura VARCHAR(100) NOT NULL,
        nota DECIMAL(5,2),
        periodo VARCHAR(20),
        observacion TEXT,
        FOREIGN KEY (id_estudiante) REFERENCES estudiantes(id_estudiante) ON DELETE CASCADE
    );
    
    CREATE TABLE IF NOT EXISTS reportes (
        id_reporte INT AUTO_INCREMENT PRIMARY KEY,
        titulo VARCHAR(200) NOT NULL,
        descripcion TEXT,
        tipo VARCHAR(50),
        fecha_generacion DATE NOT NULL,
        generado_por INT NOT NULL,
        FOREIGN KEY (generado_por) REFERENCES usuarios(id_usuario) ON DELETE CASCADE
    );
    
    CREATE TABLE IF NOT EXISTS observaciones (
        id_observacion INT AUTO_INCREMENT PRIMARY KEY,
        id_usuario INT NOT NULL,
        tipo VARCHAR(50),
        descripcion TEXT,
        fecha DATETIME NOT NULL,
        FOREIGN KEY (id_usuario) REFERENCES usuarios(id_usuario) ON DELETE CASCADE
    );
    
    CREATE TABLE IF NOT EXISTS auditoria (
        id_auditoria INT AUTO_INCREMENT PRIMARY KEY,
        id_usuario INT NOT NULL,
        usuario VARCHAR(50),
        rol VARCHAR(20),
        accion VARCHAR(100) NOT NULL,
        modulo VARCHAR(50),
        fecha DATETIME NOT NULL,
        ip VARCHAR(45),
        FOREIGN KEY (id_usuario) REFERENCES usuarios(id_usuario) ON DELETE CASCADE
    );
    
    CREATE TABLE IF NOT EXISTS detalle_auditoria (
        id_detalle INT AUTO_INCREMENT PRIMARY KEY,
        id_auditoria INT NOT NULL,
        campo VARCHAR(50),
        valor_anterior TEXT,
        valor_nuevo TEXT,
        FOREIGN KEY (id_auditoria) REFERENCES auditoria(id_auditoria) ON DELETE CASCADE
    );
    
    CREATE TABLE IF NOT EXISTS reset_codigos (
        id INT AUTO_INCREMENT PRIMARY KEY,
        id_usuario INT NOT NULL,
        codigo VARCHAR(6) NOT NULL,
        fecha_expiracion DATETIME NOT NULL,
        usado BOOLEAN DEFAULT FALSE,
        FOREIGN KEY (id_usuario) REFERENCES usuarios(id_usuario) ON DELETE CASCADE
    );
    """
    
    try:
        commands = sql_script.split(';')
        for command in commands:
            command = command.strip()
            if command:
                cursor.execute(command)
        conexion.commit()
        print("✅ Tablas creadas exitosamente")
    except Exception as e:
        print(f"❌ Error al crear tablas: {e}")
    finally:
        cursor.close()
        conexion.close()

# ==========================
# FUNCIÓN PARA INICIALIZAR USUARIOS
# ==========================
def inicializar_usuarios():
    conexion = obtener_conexion()
    if conexion is None:
        print("⚠️ No se pudo conectar a la BD para inicializar usuarios.")
        return
    cursor = conexion.cursor()
    try:
        cursor.execute("SELECT COUNT(*) AS total FROM usuarios WHERE rol = 'administrador'")
        if cursor.fetchone()["total"] == 0:
            password_hash = generate_password_hash("admin123")
            cursor.execute(
                "INSERT INTO usuarios (usuario, password, rol, email) VALUES (%s, %s, %s, %s)",
                ("admin", password_hash, "administrador", "admin@uneti.edu.ve")
            )
            print("✅ Usuario administrador creado: admin / admin123")
        cursor.execute("SELECT COUNT(*) AS total FROM usuarios WHERE rol = 'docente'")
        if cursor.fetchone()["total"] == 0:
            password_hash = generate_password_hash("docente123")
            cursor.execute(
                "INSERT INTO usuarios (usuario, password, rol, email) VALUES (%s, %s, %s, %s)",
                ("docente", password_hash, "docente", "docente@uneti.edu.ve")
            )
            print("✅ Usuario docente creado: docente / docente123")
        cursor.execute("SELECT COUNT(*) AS total FROM usuarios WHERE rol = 'estudiante'")
        if cursor.fetchone()["total"] == 0:
            password_hash = generate_password_hash("estudiante123")
            cursor.execute(
                "INSERT INTO usuarios (usuario, password, rol, email) VALUES (%s, %s, %s, %s)",
                ("estudiante", password_hash, "estudiante", "estudiante@uneti.edu.ve")
            )
            print("✅ Usuario estudiante creado: estudiante / estudiante123")
        conexion.commit()
    except Exception as e:
        print(f"❌ Error al crear usuarios por defecto: {e}")
    finally:
        cursor.close()
        conexion.close()

# ============================================
# RUTAS PÚBLICAS
# ============================================
@app.route("/")
def inicio():
    return render_template("inicio.html")

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        usuario = request.form["usuario"]
        password = request.form["password"]

        conexion = obtener_conexion()
        if conexion is None:
            return render_template("login.html", error="Error de conexión con la base de datos.")

        cursor = conexion.cursor()
        cursor.execute("SELECT * FROM usuarios WHERE usuario=%s", (usuario,))
        usuario_db = cursor.fetchone()
        cursor.close()
        conexion.close()

        if usuario_db and check_password_hash(usuario_db["password"], password):
            session["id_usuario"] = usuario_db["id_usuario"]
            session["usuario"] = usuario_db["usuario"]
            session["rol"] = usuario_db["rol"]
            session.permanent = True

            ip = request.remote_addr
            fecha_hora = obtener_fecha_venezuela()
            dia_semana = fecha_hora.strftime("%A")
            fecha_formateada = fecha_hora.strftime("%d/%m/%Y")
            hora_formateada = fecha_hora.strftime("%I:%M:%S %p")

            registrar_auditoria("Inicio de sesión", "Login")

            html = f"""
            <h2 style="color:#003366;">🔐 Inicio de sesión en SIS-UNETI</h2>
            <p><strong>Usuario:</strong> {session['usuario']}</p>
            <p><strong>Rol:</strong> {session['rol']}</p>
            <p><strong>Día:</strong> {dia_semana}</p>
            <p><strong>Fecha:</strong> {fecha_formateada}</p>
            <p><strong>Hora:</strong> {hora_formateada}</p>
            <p><strong>Dirección IP:</strong> {ip}</p>
            <hr>
            <p>Este correo es una notificación automática del sistema SIS-UNETI.</p>
            """
            enviar_correo(
                "docoutofernandodaniel@gmail.com",
                f"🔐 Inicio de sesión - {session['usuario']} - {fecha_formateada}",
                html
            )

            if usuario_db["rol"] == "administrador":
                return redirect("/administrador")
            elif usuario_db["rol"] == "docente":
                return redirect("/docente")
            else:
                return redirect("/estudiante")
        else:
            return render_template("login.html", error="Usuario o contraseña incorrectos")

    return render_template("login.html")

# ============================================
# RECUPERACIÓN DE CONTRASEÑA
# ============================================
def generar_codigo_recuperacion(id_usuario):
    codigo = str(random.randint(100000, 999999))
    expiracion = datetime.now() + timedelta(minutes=10)
    conexion = obtener_conexion()
    if conexion is None:
        return None
    cursor = conexion.cursor()
    try:
        cursor.execute("DELETE FROM reset_codigos WHERE id_usuario = %s", (id_usuario,))
        cursor.execute("""
            INSERT INTO reset_codigos (id_usuario, codigo, fecha_expiracion)
            VALUES (%s, %s, %s)
        """, (id_usuario, codigo, expiracion))
        conexion.commit()
        return codigo
    except Exception as e:
        print(f"❌ Error al generar código: {e}")
        return None
    finally:
        cursor.close()
        conexion.close()

def verificar_codigo_bd(id_usuario, codigo):
    conexion = obtener_conexion()
    if conexion is None:
        return False
    cursor = conexion.cursor()
    try:
        cursor.execute("""
            SELECT id, fecha_expiracion, usado
            FROM reset_codigos
            WHERE id_usuario = %s AND codigo = %s
        """, (id_usuario, codigo))
        resultado = cursor.fetchone()
        if not resultado:
            return False
        if resultado["usado"] or datetime.now() > resultado["fecha_expiracion"]:
            return False
        return True
    except Exception as e:
        print(f"❌ Error al verificar código: {e}")
        return False
    finally:
        cursor.close()
        conexion.close()

def marcar_codigo_usado_bd(id_usuario, codigo):
    conexion = obtener_conexion()
    if conexion is None:
        return False
    cursor = conexion.cursor()
    try:
        cursor.execute("""
            UPDATE reset_codigos
            SET usado = TRUE
            WHERE id_usuario = %s AND codigo = %s
        """, (id_usuario, codigo))
        conexion.commit()
        return True
    except Exception as e:
        print(f"❌ Error al marcar código: {e}")
        return False
    finally:
        cursor.close()
        conexion.close()

@app.route("/recuperar", methods=["GET", "POST"])
def recuperar():
    if request.method == "POST":
        usuario = request.form.get("usuario")
        if not usuario:
            return render_template("recuperar.html", error="Por favor, ingresa tu usuario.")
        conexion = obtener_conexion()
        if conexion is None:
            return render_template("recuperar.html", error="Error de conexión con la base de datos.")
        cursor = conexion.cursor()
        cursor.execute("SELECT id_usuario, usuario, email FROM usuarios WHERE usuario = %s", (usuario,))
        usuario_db = cursor.fetchone()
        cursor.close()
        conexion.close()
        if not usuario_db:
            return render_template("recuperar.html", error="Usuario no encontrado.")
        codigo = generar_codigo_recuperacion(usuario_db["id_usuario"])
        if not codigo:
            return render_template("recuperar.html", error="Error al generar el código de recuperación.")
        html = f"""
        <h2 style="color:#003366;">🔐 Código de recuperación - SIS-UNETI</h2>
        <p>Hola <strong>{usuario_db["usuario"]}</strong>,</p>
        <p>Has solicitado restablecer tu contraseña.</p>
        <p><strong>Tu código de recuperación es:</strong></p>
        <div style="font-size:32px; font-weight:bold; background:#F0F0F0; padding:20px; text-align:center; border-radius:10px; margin:20px 0; letter-spacing:5px;">
            {codigo}
        </div>
        <p>Ingresa este código en la página de verificación para crear una nueva contraseña.</p>
        <p>Este código es válido por <strong>10 minutos</strong>.</p>
        <br>
        <p>Si no solicitaste este cambio, ignora este mensaje.</p>
        <hr>
        <p style="font-size:12px; color:#888;">SIS-UNETI - Sistema Integrado de Servicios Socioeconómicos</p>
        """
        email_destino = usuario_db.get("email", "docoutofernandodaniel@gmail.com")
        resultado = enviar_correo(email_destino, "🔐 Código de recuperación - SIS-UNETI", html)
        if resultado:
            return render_template("recuperar.html", mensaje="✅ Se ha enviado un código de 6 dígitos a tu correo electrónico.", usuario=usuario)
        else:
            return render_template("recuperar.html", error="❌ Error al enviar el correo. Intenta nuevamente.")
    return render_template("recuperar.html")

@app.route("/verificar_codigo", methods=["GET", "POST"])
def verificar_codigo():
    if request.method == "POST":
        usuario = request.form.get("usuario")
        codigo = request.form.get("codigo")
        if not usuario or not codigo:
            return render_template("verificar_codigo.html", usuario=usuario, error="Por favor, completa todos los campos.")
        if len(codigo) != 6 or not codigo.isdigit():
            return render_template("verificar_codigo.html", usuario=usuario, error="El código debe tener 6 dígitos numéricos.")
        conexion = obtener_conexion()
        if conexion is None:
            return render_template("verificar_codigo.html", usuario=usuario, error="Error de conexión con la base de datos.")
        cursor = conexion.cursor()
        cursor.execute("SELECT id_usuario FROM usuarios WHERE usuario = %s", (usuario,))
        usuario_db = cursor.fetchone()
        cursor.close()
        conexion.close()
        if not usuario_db:
            return render_template("verificar_codigo.html", usuario=usuario, error="Usuario no encontrado.")
        if verificar_codigo_bd(usuario_db["id_usuario"], codigo):
            session["reset_usuario"] = usuario
            session["reset_codigo"] = codigo
            return redirect(url_for("restablecer"))
        else:
            return render_template("verificar_codigo.html", usuario=usuario, error="❌ Código inválido o expirado. Solicita uno nuevo.")
    usuario = request.args.get("usuario", "")
    return render_template("verificar_codigo.html", usuario=usuario)

@app.route("/restablecer", methods=["GET", "POST"])
def restablecer():
    if "reset_usuario" not in session or "reset_codigo" not in session:
        return redirect(url_for("recuperar"))
    usuario = session["reset_usuario"]
    codigo = session["reset_codigo"]
    if request.method == "POST":
        nueva_password = request.form.get("password")
        confirmar_password = request.form.get("confirmar_password")
        if not nueva_password or len(nueva_password) < 6:
            return render_template("restablecer.html", error="La contraseña debe tener al menos 6 caracteres.")
        if nueva_password != confirmar_password:
            return render_template("restablecer.html", error="Las contraseñas no coinciden.")
        conexion = obtener_conexion()
        if conexion is None:
            return render_template("restablecer.html", error="Error de conexión con la base de datos.")
        cursor = conexion.cursor()
        try:
            cursor.execute("SELECT id_usuario FROM usuarios WHERE usuario = %s", (usuario,))
            usuario_db = cursor.fetchone()
            if not usuario_db:
                return render_template("restablecer.html", error="Usuario no encontrado.")
            password_hash = generate_password_hash(nueva_password)
            cursor.execute("""
                UPDATE usuarios
                SET password = %s
                WHERE id_usuario = %s
            """, (password_hash, usuario_db["id_usuario"]))
            conexion.commit()
            marcar_codigo_usado_bd(usuario_db["id_usuario"], codigo)
            session.pop("reset_usuario", None)
            session.pop("reset_codigo", None)
            return """
            <div style="text-align:center; margin-top:50px; max-width:500px; margin-left:auto; margin-right:auto;">
                <div style="background:white; padding:40px; border-radius:15px; box-shadow:0 10px 30px rgba(0,0,0,0.1);">
                    <h2 style="color:green;">✅ ¡Contraseña actualizada!</h2>
                    <p>Tu contraseña ha sido restablecida exitosamente.</p>
                    <br>
                    <a href="/login" style="background:#003366; color:white; padding:12px 30px; border-radius:8px; text-decoration:none; display:inline-block;">
                        🚀 Ir al inicio de sesión
                    </a>
                </div>
            </div>
            """
        except Exception as e:
            print(f"❌ Error al actualizar contraseña: {e}")
            return render_template("restablecer.html", error="Error al actualizar la contraseña.")
        finally:
            cursor.close()
            conexion.close()
    return render_template("restablecer.html")

# ============================================
# DIAGNÓSTICO DE CORREO
# ============================================
@app.route("/diagnostico_correo")
def diagnostico_correo():
    try:
        import smtplib
        import socket
        
        try:
            server = smtplib.SMTP_SSL('smtp.gmail.com', 465, timeout=10)
            server.ehlo()
            resultado = "✅ Conexión SSL (puerto 465) exitosa"
            server.close()
        except Exception as e1:
            try:
                server = smtplib.SMTP('smtp.gmail.com', 587, timeout=10)
                server.ehlo()
                server.starttls()
                server.ehlo()
                resultado = "✅ Conexión TLS (puerto 587) exitosa"
                server.close()
            except Exception as e2:
                resultado = f"❌ Ambas conexiones fallaron. SSL: {e1}, TLS: {e2}"
        
        return f"""
        <h2>🔍 Diagnóstico de correo</h2>
        <p><strong>Configuración actual:</strong></p>
        <ul>
            <li><strong>MAIL_SERVER:</strong> {app.config['MAIL_SERVER']}</li>
            <li><strong>MAIL_PORT:</strong> {app.config['MAIL_PORT']}</li>
            <li><strong>MAIL_USERNAME:</strong> {app.config['MAIL_USERNAME']}</li>
            <li><strong>MAIL_PASSWORD:</strong> {'*' * len(app.config['MAIL_PASSWORD'])}</li>
        </ul>
        <p><strong>Diagnóstico:</strong> {resultado}</p>
        <br>
        <p>Si ves errores de autenticación, genera una nueva contraseña de aplicación en Gmail.</p>
        <a href="/recuperar">Volver a recuperar</a>
        """
    except Exception as e:
        return f"❌ Error en diagnóstico: {e}"

@app.route("/enviar_prueba")
@login_requerido
def enviar_prueba():
    resultado = enviar_correo(
        "docoutofernandodaniel@gmail.com",
        "📧 Prueba de correo SIS-UNETI",
        f"""
        <h1 style="color:#003366;">¡Correo enviado desde SIS-UNETI!</h1>
        <p>Si ves esto, la configuración de correo funciona correctamente. 🚀</p>
        <hr>
        <p><strong>Fecha:</strong> {datetime.now().strftime("%d/%m/%Y %H:%M")}</p>
        <p><strong>Usuario:</strong> {session.get("usuario", "Desconocido")}</p>
        """
    )
    if resultado:
        return """
        <h2 style="color:green;">✅ Correo enviado correctamente</h2>
        <p>Revisa tu bandeja de entrada o la carpeta de spam.</p>
        <a href="/administrador">Volver al panel</a>
        """
    else:
        return """
        <h2 style="color:red;">❌ Error al enviar el correo</h2>
        <p>Revisa la consola de Flask para ver el error detallado.</p>
        <a href="/administrador">Volver al panel</a>
        """

# ==========================
# CERRAR SESIÓN
# ==========================
@app.route("/logout")
@login_requerido
def logout():
    ip = request.remote_addr
    usuario = session.get("usuario", "Desconocido")
    rol = session.get("rol", "Desconocido")
    fecha_hora = obtener_fecha_venezuela()
    dia_semana = fecha_hora.strftime("%A")
    fecha_formateada = fecha_hora.strftime("%d/%m/%Y")
    hora_formateada = fecha_hora.strftime("%I:%M:%S %p")

    registrar_auditoria("Cerró sesión", "Login")

    html = f"""
    <h2 style="color:#DC3545;">🚪 Cierre de sesión en SIS-UNETI</h2>
    <p><strong>Usuario:</strong> {usuario}</p>
    <p><strong>Rol:</strong> {rol}</p>
    <p><strong>Día:</strong> {dia_semana}</p>
    <p><strong>Fecha:</strong> {fecha_formateada}</p>
    <p><strong>Hora:</strong> {hora_formateada}</p>
    <p><strong>Dirección IP:</strong> {ip}</p>
    <hr>
    <p>Este correo es una notificación automática del sistema SIS-UNETI.</p>
    """
    enviar_correo(
        "docoutofernandodaniel@gmail.com",
        f"🚪 Cierre de sesión - {usuario} - {fecha_formateada}",
        html
    )

    session.clear()
    return redirect("/login")

# ============================================
# PANELES
# ============================================
@app.route("/estudiante")
@login_requerido
@rol_requerido("estudiante")
def estudiante():
    return render_template("panel_estudiante.html")

@app.route("/docente")
@login_requerido
@rol_requerido("docente")
def docente():
    return render_template("panel_docente.html")

@app.route("/administrador")
@login_requerido
@rol_requerido("administrador")
def administrador():
    conexion = obtener_conexion()
    if conexion is None:
        return "Error de conexión con la base de datos."
    cursor = conexion.cursor()
    
    cursor.execute("SELECT COUNT(*) AS total FROM usuarios WHERE rol='estudiante'")
    total_estudiantes = cursor.fetchone()["total"]
    cursor.execute("SELECT COUNT(*) AS total FROM usuarios WHERE rol='docente'")
    total_docentes = cursor.fetchone()["total"]
    cursor.execute("SELECT COUNT(*) AS total FROM usuarios WHERE rol='administrador'")
    total_administradores = cursor.fetchone()["total"]
    cursor.execute("SELECT COUNT(*) AS total FROM solicitudes")
    total_solicitudes = cursor.fetchone()["total"]
    cursor.execute("SELECT COUNT(*) AS total FROM solicitudes WHERE estado='Pendiente'")
    pendientes = cursor.fetchone()["total"]
    cursor.execute("SELECT COUNT(*) AS total FROM solicitudes WHERE estado='Aprobada'")
    aprobadas = cursor.fetchone()["total"]
    cursor.execute("SELECT COUNT(*) AS total FROM solicitudes WHERE estado='Rechazada'")
    rechazadas = cursor.fetchone()["total"]
    
    cursor.execute("""
        SELECT 
            DATE_FORMAT(fecha, '%d/%m/%Y %H:%i') AS fecha_formateada,
            usuario,
            accion,
            modulo,
            ip
        FROM auditoria
        ORDER BY fecha DESC
        LIMIT 10
    """)
    actividad = cursor.fetchall()
    
    notificaciones = []
    if pendientes > 0:
        notificaciones.append({
            "tipo": "warning",
            "mensaje": f"⚠️ Existen {pendientes} solicitudes pendientes por revisar."
        })
    if total_solicitudes == 0:
        notificaciones.append({
            "tipo": "info",
            "mensaje": "ℹ️ No existen solicitudes registradas en el sistema."
        })
    else:
        notificaciones.append({
            "tipo": "success",
            "mensaje": "✅ El sistema funciona correctamente."
        })
    
    total_usuarios = total_estudiantes + total_docentes + total_administradores
    notificaciones.append({
        "tipo": "primary",
        "mensaje": f"👥 Usuarios registrados en el sistema: {total_usuarios}"
    })
    
    cursor.close()
    conexion.close()
    
    return render_template(
        "panel_administrador.html",
        estudiantes=total_estudiantes,
        docentes=total_docentes,
        administradores=total_administradores,
        solicitudes=total_solicitudes,
        pendientes=pendientes,
        aprobadas=aprobadas,
        rechazadas=rechazadas,
        actividad=actividad,
        notificaciones=notificaciones
    )

# ============================================
# GESTIÓN DE USUARIOS, DOCENTES, ESTUDIANTES
# ============================================
@app.route("/usuarios")
@login_requerido
@rol_requerido("administrador")
def usuarios():
    conexion = obtener_conexion()
    if conexion is None:
        return "Error de conexión con la base de datos."
    cursor = conexion.cursor()
    cursor.execute("SELECT * FROM usuarios ORDER BY id_usuario DESC")
    datos = cursor.fetchall()
    cursor.close()
    conexion.close()
    return render_template("usuarios.html", usuarios=datos)

@app.route("/docentes")
@login_requerido
@rol_requerido("administrador")
def docentes():
    conexion = obtener_conexion()
    if conexion is None:
        return "Error de conexión con la base de datos."
    cursor = conexion.cursor()
    cursor.execute("SELECT * FROM docentes ORDER BY id_docente DESC")
    datos = cursor.fetchall()
    cursor.close()
    conexion.close()
    return render_template("docentes.html", docentes=datos)

@app.route("/estudiantes")
@login_requerido
@rol_requerido("administrador")
def estudiantes():
    conexion = obtener_conexion()
    if conexion is None:
        return "Error de conexión con la base de datos."
    cursor = conexion.cursor()
    cursor.execute("SELECT * FROM estudiantes ORDER BY id_estudiante DESC")
    datos = cursor.fetchall()
    cursor.close()
    conexion.close()
    return render_template("estudiantes.html", estudiantes=datos)

# ============================================
# GESTIÓN DE SOLICITUDES
# ============================================
@app.route("/solicitudes")
@login_requerido
def solicitudes():
    if session.get("rol") not in ["administrador", "docente"]:
        return "⛔ Acceso no autorizado. Solo administradores y docentes pueden ver las solicitudes.", 403
    
    conexion = obtener_conexion()
    if conexion is None:
        return "Error de conexión con la base de datos."
    cursor = conexion.cursor()
    cursor.execute("""
        SELECT solicitudes.*, usuarios.usuario
        FROM solicitudes
        INNER JOIN usuarios ON solicitudes.id_usuario = usuarios.id_usuario
        ORDER BY solicitudes.id_solicitud DESC
    """)
    datos = cursor.fetchall()
    cursor.close()
    conexion.close()
    return render_template("solicitudes.html", solicitudes=datos)

@app.route("/mis_solicitudes")
@login_requerido
def mis_solicitudes():
    conexion = obtener_conexion()
    if conexion is None:
        return "Error de conexión con la base de datos."
    cursor = conexion.cursor()
    cursor.execute("SELECT * FROM solicitudes WHERE id_usuario=%s ORDER BY id_solicitud DESC", (session["id_usuario"],))
    datos = cursor.fetchall()
    cursor.close()
    conexion.close()
    return render_template("mis_solicitudes.html", solicitudes=datos)

@app.route("/nueva_solicitud", methods=["GET", "POST"])
@login_requerido
def nueva_solicitud():
    if session.get("rol") != "estudiante":
        return render_template("error.html", mensaje="⛔ Acceso denegado. Solo los estudiantes pueden crear solicitudes."), 403
    if request.method == "POST":
        fecha = request.form["fecha"]
        tipo = request.form["tipo_solicitud"]
        descripcion = request.form["descripcion"]
        conexion = obtener_conexion()
        if conexion is None:
            return "Error de conexión con la base de datos."
        cursor = conexion.cursor()
        cursor.execute("""
            INSERT INTO solicitudes (fecha, tipo_solicitud, descripcion, estado, id_usuario)
            VALUES (%s, %s, %s, %s, %s)
        """, (fecha, tipo, descripcion, "Pendiente", session["id_usuario"]))
        conexion.commit()
        registrar_auditoria("Creó una solicitud", "Solicitudes")
        cursor.close()
        conexion.close()
        html = f"""
        <h2>✅ Solicitud registrada en SIS-UNETI</h2>
        <p>Hola <strong>{session['usuario']}</strong>,</p>
        <p>Tu solicitud ha sido registrada exitosamente.</p>
        <p><strong>Detalles:</strong></p>
        <ul>
            <li><strong>Fecha:</strong> {fecha}</li>
            <li><strong>Tipo:</strong> {tipo}</li>
            <li><strong>Descripción:</strong> {descripcion}</li>
            <li><strong>Estado:</strong> Pendiente</li>
        </ul>
        <p>Recibirás una notificación cuando sea procesada.</p>
        <br>
        <p>Saludos,<br><strong>SIS-UNETI</strong></p>
        """
        enviar_correo("docoutofernandodaniel@gmail.com", "✅ Confirmación de solicitud - SIS-UNETI", html)
        return redirect("/solicitudes")
    return render_template("nueva_solicitud.html")

@app.route("/editar_solicitud/<int:id>", methods=["GET", "POST"])
@login_requerido
@rol_requerido("administrador", "docente", "estudiante")
def editar_solicitud(id):
    conexion = obtener_conexion()
    if conexion is None:
        return "Error de conexión con la base de datos."
    cursor = conexion.cursor()
    if request.method == "POST":
        tipo = request.form["tipo_solicitud"]
        descripcion = request.form["descripcion"]
        estado = request.form["estado"]
        cursor.execute("""
            UPDATE solicitudes
            SET tipo_solicitud=%s, descripcion=%s, estado=%s
            WHERE id_solicitud=%s
        """, (tipo, descripcion, estado, id))
        conexion.commit()
        registrar_auditoria("Editó una solicitud", "Solicitudes")
        cursor.close()
        conexion.close()
        return redirect("/solicitudes")
    cursor.execute("SELECT * FROM solicitudes WHERE id_solicitud=%s", (id,))
    solicitud = cursor.fetchone()
    cursor.close()
    conexion.close()
    return render_template("editar_solicitud.html", solicitud=solicitud)

@app.route("/eliminar_solicitud/<int:id>")
@login_requerido
@rol_requerido("administrador", "docente")
def eliminar_solicitud(id):
    conexion = obtener_conexion()
    if conexion is None:
        return "Error de conexión con la base de datos."
    cursor = conexion.cursor()
    cursor.execute("DELETE FROM solicitudes WHERE id_solicitud=%s", (id,))
    conexion.commit()
    registrar_auditoria("Eliminó una solicitud", "Solicitudes")
    cursor.close()
    conexion.close()
    return redirect("/solicitudes")

# ============================================
# GESTIÓN DE BENEFICIOS
# ============================================
@app.route("/beneficios")
@login_requerido
@rol_requerido("administrador", "docente", "estudiante")
def beneficios():
    conexion = obtener_conexion()
    if conexion is None:
        return "Error de conexión con la base de datos."
    cursor = conexion.cursor()
    cursor.execute("""
        SELECT beneficios.*, estudiantes.nombres, estudiantes.apellidos
        FROM beneficios
        INNER JOIN estudiantes ON beneficios.id_estudiante = estudiantes.id_estudiante
        ORDER BY beneficios.id_beneficio DESC
    """)
    datos = cursor.fetchall()
    cursor.close()
    conexion.close()
    return render_template("beneficios.html", beneficios=datos)

@app.route("/nuevo_beneficio", methods=["GET", "POST"])
@login_requerido
@rol_requerido("administrador", "docente")
def nuevo_beneficio():
    conexion = obtener_conexion()
    if conexion is None:
        return "Error de conexión con la base de datos."
    cursor = conexion.cursor()
    if request.method == "POST":
        id_estudiante = request.form["id_estudiante"]
        nombre_beneficio = request.form["nombre_beneficio"]
        descripcion = request.form["descripcion"]
        fecha_solicitud = request.form["fecha_solicitud"]
        estado = request.form["estado"]
        cursor.execute("""
            INSERT INTO beneficios (id_estudiante, nombre_beneficio, descripcion, fecha_solicitud, estado)
            VALUES (%s, %s, %s, %s, %s)
        """, (id_estudiante, nombre_beneficio, descripcion, fecha_solicitud, estado))
        conexion.commit()
        registrar_auditoria("Creó un beneficio", "Beneficios")
        cursor.close()
        conexion.close()
        return redirect("/beneficios")
    cursor.execute("SELECT id_estudiante, nombres, apellidos FROM estudiantes ORDER BY nombres ASC")
    estudiantes = cursor.fetchall()
    cursor.close()
    conexion.close()
    return render_template("nuevo_beneficio.html", estudiantes=estudiantes)

@app.route("/editar_beneficio/<int:id>", methods=["GET", "POST"])
@login_requerido
@rol_requerido("administrador", "docente")
def editar_beneficio(id):
    conexion = obtener_conexion()
    if conexion is None:
        return "Error de conexión con la base de datos."
    cursor = conexion.cursor()
    cursor.execute("SELECT * FROM beneficios WHERE id_beneficio=%s", (id,))
    beneficio_anterior = cursor.fetchone()
    if not beneficio_anterior:
        cursor.close()
        conexion.close()
        return "Beneficio no encontrado.", 404
    if request.method == "POST":
        nombre_beneficio = request.form["nombre_beneficio"]
        descripcion = request.form["descripcion"]
        estado = request.form["estado"]
        id_auditoria = registrar_auditoria("Editó un beneficio", "Beneficios")
        if nombre_beneficio != beneficio_anterior["nombre_beneficio"]:
            registrar_detalle_auditoria(
                id_auditoria,
                "nombre_beneficio",
                beneficio_anterior["nombre_beneficio"],
                nombre_beneficio
            )
        if descripcion != beneficio_anterior["descripcion"]:
            registrar_detalle_auditoria(
                id_auditoria,
                "descripcion",
                beneficio_anterior["descripcion"],
                descripcion
            )
        if estado != beneficio_anterior["estado"]:
            registrar_detalle_auditoria(
                id_auditoria,
                "estado",
                beneficio_anterior["estado"],
                estado
            )
        cursor.execute("""
            UPDATE beneficios
            SET nombre_beneficio=%s, descripcion=%s, estado=%s
            WHERE id_beneficio=%s
        """, (nombre_beneficio, descripcion, estado, id))
        conexion.commit()
        cursor.close()
        conexion.close()
        return redirect("/beneficios")
    cursor.close()
    conexion.close()
    return render_template("editar_beneficio.html", beneficio=beneficio_anterior)
    
@app.route("/eliminar_beneficio/<int:id>")
@login_requerido
@rol_requerido("administrador", "docente")
def eliminar_beneficio(id):
    conexion = obtener_conexion()
    if conexion is None:
        return "Error de conexión con la base de datos."
    cursor = conexion.cursor()
    cursor.execute("DELETE FROM beneficios WHERE id_beneficio=%s", (id,))
    conexion.commit()
    registrar_auditoria("Eliminó un beneficio", "Beneficios")
    cursor.close()
    conexion.close()
    return redirect("/beneficios")

# ============================================
# GESTIÓN DE REPORTES
# ============================================
@app.route("/reportes")
@login_requerido
@rol_requerido("administrador", "docente", "estudiante")
def reportes():
    conexion = obtener_conexion()
    if conexion is None:
        return "Error de conexión con la base de datos."
    cursor = conexion.cursor()
    cursor.execute("""
        SELECT reportes.*, usuarios.usuario
        FROM reportes
        INNER JOIN usuarios ON reportes.generado_por = usuarios.id_usuario
        ORDER BY reportes.id_reporte DESC
    """)
    datos = cursor.fetchall()
    cursor.close()
    conexion.close()
    return render_template("reportes.html", reportes=datos)

@app.route("/nuevo_reporte", methods=["GET", "POST"])
@login_requerido
@rol_requerido("administrador", "docente")
def nuevo_reporte():
    if request.method == "POST":
        titulo = request.form.get("titulo")
        descripcion = request.form.get("descripcion")
        tipo = request.form.get("tipo")
        fecha_generacion = request.form.get("fecha_generacion")
        if not titulo or not descripcion or not tipo or not fecha_generacion:
            return render_template("nuevo_reporte.html", error="Todos los campos son obligatorios.")
        conexion = obtener_conexion()
        if conexion is None:
            return "Error de conexión con la base de datos."
        cursor = conexion.cursor()
        cursor.execute("""
            INSERT INTO reportes (titulo, descripcion, tipo, fecha_generacion, generado_por)
            VALUES (%s, %s, %s, %s, %s)
        """, (titulo, descripcion, tipo, fecha_generacion, session["id_usuario"]))
        conexion.commit()
        registrar_auditoria("Generó un reporte", "Reportes")
        cursor.close()
        conexion.close()
        return redirect("/reportes")
    return render_template("nuevo_reporte.html")

@app.route("/editar_reporte/<int:id>", methods=["GET", "POST"])
@login_requerido
@rol_requerido("administrador", "docente")
def editar_reporte(id):
    conexion = obtener_conexion()
    if conexion is None:
        return "Error de conexión con la base de datos."
    cursor = conexion.cursor()
    cursor.execute("SELECT * FROM reportes WHERE id_reporte=%s", (id,))
    reporte = cursor.fetchone()
    cursor.close()
    if not reporte:
        conexion.close()
        return "Reporte no encontrado.", 404
    if request.method == "POST":
        titulo = request.form.get("titulo")
        descripcion = request.form.get("descripcion")
        tipo = request.form.get("tipo")
        fecha_generacion = request.form.get("fecha_generacion")
        if not titulo or not descripcion or not tipo or not fecha_generacion:
            return render_template("editar_reporte.html", reporte=reporte, error="Todos los campos son obligatorios.")
        cursor = conexion.cursor()
        try:
            cursor.execute("""
                UPDATE reportes
                SET titulo=%s, descripcion=%s, tipo=%s, fecha_generacion=%s
                WHERE id_reporte=%s
            """, (titulo, descripcion, tipo, fecha_generacion, id))
            conexion.commit()
            registrar_auditoria("Editó un reporte", "Reportes")
            cursor.close()
            conexion.close()
            return redirect("/reportes")
        except pymysql.OperationalError as e:
            cursor.close()
            conexion.close()
            return render_template("editar_reporte.html", reporte=reporte, error=f"Error al actualizar: {str(e)}")
    conexion.close()
    return render_template("editar_reporte.html", reporte=reporte)

@app.route("/eliminar_reporte/<int:id>")
@login_requerido
@rol_requerido("administrador", "docente")
def eliminar_reporte(id):
    conexion = obtener_conexion()
    if conexion is None:
        return "Error de conexión con la base de datos."
    cursor = conexion.cursor()
    cursor.execute("DELETE FROM reportes WHERE id_reporte=%s", (id,))
    conexion.commit()
    registrar_auditoria("Eliminó un reporte", "Reportes")
    cursor.close()
    conexion.close()
    return redirect("/reportes")

# ============================================
# GESTIÓN DE EVALUACIONES
# ============================================
@app.route("/evaluaciones")
@login_requerido
@rol_requerido("administrador", "docente", "estudiante")
def evaluaciones():
    conexion = obtener_conexion()
    if conexion is None:
        return "Error de conexión con la base de datos."
    cursor = conexion.cursor()
    cursor.execute("""
        SELECT evaluaciones.*, estudiantes.nombres, estudiantes.apellidos
        FROM evaluaciones
        INNER JOIN estudiantes ON evaluaciones.id_estudiante = estudiantes.id_estudiante
        ORDER BY evaluaciones.id_evaluacion DESC
    """)
    datos = cursor.fetchall()
    cursor.close()
    conexion.close()
    return render_template("evaluaciones.html", evaluaciones=datos)

@app.route("/nueva_evaluacion", methods=["GET", "POST"])
@login_requerido
@rol_requerido("administrador", "docente")
def nueva_evaluacion():
    conexion = obtener_conexion()
    if conexion is None:
        return "Error de conexión con la base de datos."
    cursor = conexion.cursor()
    if request.method == "POST":
        id_estudiante = request.form["id_estudiante"]
        asignatura = request.form["asignatura"]
        nota = request.form["nota"]
        periodo = request.form["periodo"]
        observacion = request.form["observacion"]
        cursor.execute("""
            INSERT INTO evaluaciones (id_estudiante, asignatura, nota, periodo, observacion)
            VALUES (%s, %s, %s, %s, %s)
        """, (id_estudiante, asignatura, nota, periodo, observacion))
        conexion.commit()
        registrar_auditoria("Registró una evaluación", "Evaluaciones")
        cursor.close()
        conexion.close()
        return redirect("/evaluaciones")
    cursor.execute("SELECT id_estudiante, nombres, apellidos FROM estudiantes ORDER BY nombres ASC")
    estudiantes = cursor.fetchall()
    cursor.close()
    conexion.close()
    return render_template("nueva_evaluacion.html", estudiantes=estudiantes)

@app.route("/editar_evaluacion/<int:id>", methods=["GET", "POST"])
@login_requerido
@rol_requerido("administrador", "docente")
def editar_evaluacion(id):
    conexion = obtener_conexion()
    if conexion is None:
        return "Error de conexión con la base de datos."
    cursor = conexion.cursor()
    if request.method == "POST":
        asignatura = request.form["asignatura"]
        nota = request.form["nota"]
        periodo = request.form["periodo"]
        observacion = request.form["observacion"]
        cursor.execute("""
            UPDATE evaluaciones
            SET asignatura=%s, nota=%s, periodo=%s, observacion=%s
            WHERE id_evaluacion=%s
        """, (asignatura, nota, periodo, observacion, id))
        conexion.commit()
        registrar_auditoria("Editó una evaluación", "Evaluaciones")
        cursor.close()
        conexion.close()
        return redirect("/evaluaciones")
    cursor.execute("SELECT * FROM evaluaciones WHERE id_evaluacion=%s", (id,))
    evaluacion = cursor.fetchone()
    cursor.close()
    conexion.close()
    return render_template("editar_evaluacion.html", evaluacion=evaluacion)

@app.route("/eliminar_evaluacion/<int:id>")
@login_requerido
@rol_requerido("administrador", "docente")
def eliminar_evaluacion(id):
    conexion = obtener_conexion()
    if conexion is None:
        return "Error de conexión con la base de datos."
    cursor = conexion.cursor()
    cursor.execute("DELETE FROM evaluaciones WHERE id_evaluacion=%s", (id,))
    conexion.commit()
    registrar_auditoria("Eliminó una evaluación", "Evaluaciones")
    cursor.close()
    conexion.close()
    return redirect("/evaluaciones")

# ============================================
# GESTIÓN DE OBSERVACIONES
# ============================================
@app.route("/observaciones")
@login_requerido
@rol_requerido("administrador", "docente")
def observaciones():
    conexion = obtener_conexion()
    if conexion is None:
        return "Error de conexión con la base de datos."
    cursor = conexion.cursor()
    cursor.execute("""
        SELECT observaciones.*, usuarios.usuario
        FROM observaciones
        INNER JOIN usuarios ON observaciones.id_usuario = usuarios.id_usuario
        ORDER BY observaciones.id_observacion DESC
    """)
    datos = cursor.fetchall()
    cursor.close()
    conexion.close()
    return render_template("observaciones.html", observaciones=datos)

@app.route("/nueva_observacion", methods=["GET", "POST"])
@login_requerido
@rol_requerido("administrador", "docente", "estudiante")
def nueva_observacion():
    if request.method == "POST":
        tipo = request.form["tipo"]
        descripcion = request.form["descripcion"]
        conexion = obtener_conexion()
        if conexion is None:
            return "Error de conexión con la base de datos."
        cursor = conexion.cursor()
        cursor.execute("""
            INSERT INTO observaciones (id_usuario, tipo, descripcion, fecha)
            VALUES (%s, %s, %s, %s)
        """, (session["id_usuario"], tipo, descripcion, obtener_fecha_venezuela()))
        conexion.commit()
        registrar_auditoria("Registró una observación", "Observaciones")
        cursor.close()
        conexion.close()
        if session["rol"] in ["administrador", "docente"]:
            return redirect("/observaciones")
        else:
            return redirect("/mis_observaciones")
    return render_template("nueva_observacion.html")

@app.route("/mis_observaciones")
@login_requerido
@rol_requerido("estudiante")
def mis_observaciones():
    conexion = obtener_conexion()
    if conexion is None:
        return "Error de conexión con la base de datos."
    cursor = conexion.cursor()
    cursor.execute("SELECT * FROM observaciones WHERE id_usuario=%s ORDER BY id_observacion DESC", (session["id_usuario"],))
    datos = cursor.fetchall()
    cursor.close()
    conexion.close()
    return render_template("mis_observaciones.html", observaciones=datos)

@app.route("/editar_observacion/<int:id>", methods=["GET", "POST"])
@login_requerido
@rol_requerido("administrador", "docente")
def editar_observacion(id):
    conexion = obtener_conexion()
    if conexion is None:
        return "Error de conexión con la base de datos."
    cursor = conexion.cursor()
    if request.method == "POST":
        tipo = request.form["tipo"]
        descripcion = request.form["descripcion"]
        cursor.execute("""
            UPDATE observaciones
            SET tipo=%s, descripcion=%s
            WHERE id_observacion=%s
        """, (tipo, descripcion, id))
        conexion.commit()
        registrar_auditoria("Editó una observación", "Observaciones")
        cursor.close()
        conexion.close()
        return redirect("/observaciones")
    cursor.execute("SELECT * FROM observaciones WHERE id_observacion=%s", (id,))
    observacion = cursor.fetchone()
    cursor.close()
    conexion.close()
    return render_template("editar_observacion.html", observacion=observacion)

@app.route("/eliminar_observacion/<int:id>")
@login_requerido
@rol_requerido("administrador", "docente")
def eliminar_observacion(id):
    conexion = obtener_conexion()
    if conexion is None:
        return "Error de conexión con la base de datos."
    cursor = conexion.cursor()
    cursor.execute("DELETE FROM observaciones WHERE id_observacion=%s", (id,))
    conexion.commit()
    registrar_auditoria("Eliminó una observación", "Observaciones")
    cursor.close()
    conexion.close()
    return redirect("/observaciones")

# ============================================
# AUDITORÍA DEL SISTEMA
# ============================================
@app.route("/auditoria")
@login_requerido
@rol_requerido("administrador")
def auditoria():
    pagina = request.args.get("pagina", 1, type=int)
    por_pagina = 10
    offset = (pagina - 1) * por_pagina

    buscar = request.args.get("buscar", "").strip()
    rol = request.args.get("rol", "").strip()
    modulo = request.args.get("modulo", "").strip()
    fecha_desde = request.args.get("desde", "").strip()
    fecha_hasta = request.args.get("hasta", "").strip()

    conexion = obtener_conexion()
    if conexion is None:
        return "Error de conexión con la base de datos."

    cursor = conexion.cursor()

    cursor.execute("SELECT COUNT(*) AS total FROM auditoria")
    total_acciones = cursor.fetchone()["total"]

    cursor.execute("SELECT COUNT(*) AS total FROM auditoria WHERE accion='Inicio de sesión'")
    total_logins = cursor.fetchone()["total"]

    cursor.execute("SELECT COUNT(*) AS total FROM auditoria WHERE accion LIKE 'Editó%'")
    total_ediciones = cursor.fetchone()["total"]

    cursor.execute("SELECT COUNT(*) AS total FROM auditoria WHERE accion LIKE 'Eliminó%'")
    total_eliminaciones = cursor.fetchone()["total"]

    cursor.execute("SELECT usuario, COUNT(*) AS total FROM auditoria GROUP BY usuario ORDER BY total DESC LIMIT 1")
    usuario_activo = cursor.fetchone()

    cursor.execute("SELECT modulo, COUNT(*) AS total FROM auditoria GROUP BY modulo ORDER BY total DESC LIMIT 1")
    modulo_principal = cursor.fetchone()

    cursor.execute("SELECT rol, COUNT(*) AS total FROM auditoria WHERE accion='Inicio de sesión' GROUP BY rol")
    accesos_roles = cursor.fetchall()

    cursor.execute("SELECT DATE(fecha) AS dia, COUNT(*) AS total FROM auditoria GROUP BY DATE(fecha) ORDER BY dia DESC")
    actividad_diaria = cursor.fetchall()

    cursor.execute("SELECT modulo, COUNT(*) AS total FROM auditoria GROUP BY modulo ORDER BY total DESC")
    acciones_modulos = cursor.fetchall()

    cursor.execute("SELECT usuario, COUNT(*) AS total FROM auditoria GROUP BY usuario ORDER BY total DESC LIMIT 5")
    usuarios_activos = cursor.fetchall()

    cursor.execute("SELECT accion, COUNT(*) AS total FROM auditoria GROUP BY accion ORDER BY total DESC LIMIT 10")
    acciones_tipo = cursor.fetchall()

    consulta = """
        SELECT id_auditoria, usuario, rol, accion, modulo, fecha, ip
        FROM auditoria
        WHERE 1=1
    """
    parametros = []

    if buscar:
        consulta += " AND usuario LIKE %s"
        parametros.append(f"%{buscar}%")
    if rol:
        consulta += " AND rol=%s"
        parametros.append(rol)
    if modulo:
        consulta += " AND modulo=%s"
        parametros.append(modulo)
    if fecha_desde:
        consulta += " AND DATE(fecha) >= %s"
        parametros.append(fecha_desde)
    if fecha_hasta:
        consulta += " AND DATE(fecha) <= %s"
        parametros.append(fecha_hasta)

    consulta_total = consulta.replace("SELECT id_auditoria, usuario, rol, accion, modulo, fecha, ip", "SELECT COUNT(*) AS total")
    cursor.execute(consulta_total, parametros)
    total_registros = cursor.fetchone()["total"]
    total_paginas = (total_registros + por_pagina - 1) // por_pagina if total_registros > 0 else 1

    consulta += " ORDER BY fecha DESC LIMIT %s OFFSET %s"
    parametros.append(por_pagina)
    parametros.append(offset)

    cursor.execute(consulta, parametros)
    auditorias = cursor.fetchall()

    cursor.close()
    conexion.close()

    return render_template(
        "auditoria.html",
        auditorias=auditorias,
        total_acciones=total_acciones,
        total_logins=total_logins,
        total_ediciones=total_ediciones,
        total_eliminaciones=total_eliminaciones,
        usuario_activo=usuario_activo,
        modulo_principal=modulo_principal,
        accesos_roles=accesos_roles,
        actividad_diaria=actividad_diaria,
        acciones_modulos=acciones_modulos,
        usuarios_activos=usuarios_activos,
        acciones_tipo=acciones_tipo,
        buscar=buscar,
        rol=rol,
        modulo=modulo,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        pagina=pagina,
        total_paginas=total_paginas
    )

# ============================================
# EXPORTAR AUDITORÍA A PDF
# ============================================
@app.route("/auditoria/pdf")
@login_requerido
@rol_requerido("administrador")
def exportar_pdf():
    try:
        conexion = obtener_conexion()
        if conexion is None:
            return "Error de conexión con la base de datos."
        cursor = conexion.cursor()
        cursor.execute("SELECT id_auditoria, usuario, rol, accion, modulo, fecha, ip FROM auditoria ORDER BY fecha DESC")
        auditorias = cursor.fetchall()
        cursor.execute("SELECT COUNT(*) AS total FROM auditoria")
        total_acciones = cursor.fetchone()["total"]
        cursor.execute("SELECT COUNT(*) AS total FROM auditoria WHERE accion='Inicio de sesión'")
        total_logins = cursor.fetchone()["total"]
        cursor.execute("SELECT COUNT(*) AS total FROM auditoria WHERE accion LIKE 'Editó%'")
        total_ediciones = cursor.fetchone()["total"]
        cursor.execute("SELECT COUNT(*) AS total FROM auditoria WHERE accion LIKE 'Eliminó%'")
        total_eliminaciones = cursor.fetchone()["total"]
        cursor.close()
        conexion.close()
        buffer = io.BytesIO()
        pdf = SimpleDocTemplate(buffer, pagesize=letter)
        elementos = []
        estilos = getSampleStyleSheet()
        elementos.append(Paragraph("SIS-UNETI<br/>Centro de Auditoría del Sistema", estilos["Title"]))
        elementos.append(Spacer(1, 20))
        fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M")
        elementos.append(Paragraph(f"Fecha de generación: {fecha_actual}", estilos["Normal"]))
        elementos.append(Spacer(1, 20))
        resumen = [
            ["Estadística", "Cantidad"],
            ["Total acciones", str(total_acciones)],
            ["Inicios de sesión", str(total_logins)],
            ["Ediciones", str(total_ediciones)],
            ["Eliminaciones", str(total_eliminaciones)]
        ]
        tabla_resumen = Table(resumen, colWidths=[200, 100])
        tabla_resumen.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('PADDING', (0, 0), (-1, -1), 6)
        ]))
        elementos.append(tabla_resumen)
        elementos.append(Spacer(1, 25))
        datos = [["ID", "Usuario", "Rol", "Acción", "Módulo", "Fecha", "IP"]]
        for reg in auditorias:
            datos.append([str(reg["id_auditoria"]), reg["usuario"], reg["rol"], reg["accion"], reg["modulo"], str(reg["fecha"]), reg["ip"]])
        tabla = Table(datos, repeatRows=1)
        tabla.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('PADDING', (0, 0), (-1, -1), 4)
        ]))
        elementos.append(tabla)
        pdf.build(elementos)
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name=f"Auditoria_SIS_UNETI_{datetime.now().strftime('%Y%m%d')}.pdf", mimetype="application/pdf")
    except Exception as e:
        return f"Error al generar el PDF: {str(e)}"

# ============================================
# EXPORTAR AUDITORÍA A EXCEL
# ============================================
@app.route("/auditoria/excel")
@login_requerido
@rol_requerido("administrador")
def exportar_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        conexion = obtener_conexion()
        if conexion is None:
            return "Error de conexión con la base de datos."
        cursor = conexion.cursor()
        buscar = request.args.get("buscar", "").strip()
        rol = request.args.get("rol", "").strip()
        modulo = request.args.get("modulo", "").strip()
        fecha_desde = request.args.get("desde", "").strip()
        fecha_hasta = request.args.get("hasta", "").strip()
        consulta = """
            SELECT id_auditoria, usuario, rol, accion, modulo, fecha, ip
            FROM auditoria
            WHERE 1=1
        """
        parametros = []
        if buscar:
            consulta += " AND usuario LIKE %s"
            parametros.append(f"%{buscar}%")
        if rol:
            consulta += " AND rol=%s"
            parametros.append(rol)
        if modulo:
            consulta += " AND modulo=%s"
            parametros.append(modulo)
        if fecha_desde:
            consulta += " AND DATE(fecha) >= %s"
            parametros.append(fecha_desde)
        if fecha_hasta:
            consulta += " AND DATE(fecha) <= %s"
            parametros.append(fecha_hasta)
        consulta += " ORDER BY fecha DESC"
        cursor.execute(consulta, parametros)
        auditorias = cursor.fetchall()
        cursor.close()
        conexion.close()
        wb = Workbook()
        ws = wb.active
        ws.title = "Auditoría SIS-UNETI"
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        headers = ["ID", "Usuario", "Rol", "Acción", "Módulo", "Fecha", "IP"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        row_num = 2
        for reg in auditorias:
            ws.cell(row=row_num, column=1, value=reg["id_auditoria"])
            ws.cell(row=row_num, column=2, value=reg["usuario"])
            ws.cell(row=row_num, column=3, value=reg["rol"])
            ws.cell(row=row_num, column=4, value=reg["accion"])
            ws.cell(row=row_num, column=5, value=reg["modulo"])
            ws.cell(row=row_num, column=6, value=str(reg["fecha"]))
            ws.cell(row=row_num, column=7, value=reg["ip"])
            row_num += 1
        for col in range(1, 8):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 18
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name=f"Auditoria_SIS_UNETI_{datetime.now().strftime('%Y%m%d')}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return f"Error al generar el Excel: {str(e)}"

# ============================================
# GESTIÓN DE HISTORIALES (LIMPIEZA)
# ============================================
@app.route("/limpiar_historiales", methods=["GET"])
@login_requerido
@rol_requerido("administrador")
def limpiar_historiales():
    conexion = obtener_conexion()
    if conexion is None:
        flash("Error de conexión con la base de datos.", "error")
        return redirect(url_for("administrador"))
    
    cursor = conexion.cursor()
    cursor.execute("SELECT COUNT(*) AS total FROM auditoria")
    total_auditoria = cursor.fetchone()["total"]
    cursor.execute("SELECT COUNT(*) AS total FROM detalle_auditoria")
    total_detalle = cursor.fetchone()["total"]
    cursor.close()
    conexion.close()
    
    return render_template(
        "limpiar_historiales.html",
        total_auditoria=total_auditoria,
        total_detalle=total_detalle
    )

@app.route("/ejecutar_limpieza", methods=["POST"])
@login_requerido
@rol_requerido("administrador")
def ejecutar_limpieza():
    opcion = request.form.get("opcion", "completa")
    
    conexion = obtener_conexion()
    if conexion is None:
        flash("Error de conexión con la base de datos.", "error")
        return redirect(url_for("limpiar_historiales"))
    
    cursor = conexion.cursor()
    
    try:
        if opcion == "completa":
            cursor.execute("DELETE FROM detalle_auditoria")
            cursor.execute("DELETE FROM auditoria")
            mensaje = "✅ Historial de auditoría eliminado COMPLETAMENTE."
        elif opcion == "antiguo":
            cursor.execute("""
                DELETE FROM detalle_auditoria 
                WHERE id_auditoria IN (
                    SELECT id_auditoria FROM auditoria 
                    WHERE fecha < DATE_SUB(NOW(), INTERVAL 180 DAY)
                )
            """)
            cursor.execute("DELETE FROM auditoria WHERE fecha < DATE_SUB(NOW(), INTERVAL 180 DAY)")
            mensaje = "✅ Historial anterior a 6 meses eliminado."
        else:
            flash("⚠️ Opción no válida.", "error")
            return redirect(url_for("limpiar_historiales"))
        
        conexion.commit()
        registrar_auditoria(f"Limpieza de historial - Opción: {opcion}", "Mantenimiento")
        cursor.close()
        conexion.close()
        
        flash(mensaje, "success")
        return redirect(url_for("limpiar_historiales"))
    
    except Exception as e:
        try:
            conexion.rollback()
        except:
            pass
        cursor.close()
        conexion.close()
        flash(f"❌ Error al limpiar: {str(e)}", "error")
        return redirect(url_for("limpiar_historiales"))

# ==========================
# MANEJO DE ERRORES
# ==========================
@app.errorhandler(404)
def pagina_no_encontrada(error):
    return render_template("error404.html"), 404

@app.errorhandler(500)
def error_servidor(error):
    return render_template("error500.html"), 500

# ==========================
# INICIALIZACIÓN AL INICIAR
# ==========================
with app.app_context():
    crear_tablas()
    inicializar_usuarios()

# ==========================
# INICIO DEL SERVIDOR
# ==========================
if __name__ == "__main__":
    print("""
    =====================================
        🚀 SIS-UNETI INICIANDO...
    =====================================
    """)
    conexion = obtener_conexion()
    if conexion:
        print("✅ Conexión MySQL establecida correctamente.")
        conexion.close()
    else:
        print("⚠️ Servidor iniciado sin conexión a Base de Datos.")
    app.run(host="0.0.0.0", port=5000, debug=True)
